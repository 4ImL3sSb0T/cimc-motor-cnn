"""
从 xlsx/csv 文件加载 IMU 加速度数据，去除直流偏移
"""

import csv
import numpy as np
import openpyxl
from pathlib import Path


def find_data_file(data_dir: str | Path) -> Path | None:
    """在目录中找到第一个 xlsx 或 csv 文件 (优先 xlsx)"""
    p = Path(data_dir)
    xlsx_files = sorted(p.glob("*.xlsx"))
    if xlsx_files:
        return xlsx_files[0]
    csv_files = sorted(p.glob("*.csv"))
    return csv_files[0] if csv_files else None


def load_xlsx(path: str | Path) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    读取 xlsx 中的 3 轴加速度数据。
    列布局: PC Date, PC Time, Line 1(X), Line 2(Y), Line 3(Z)
    返回 (ax, ay, az), dtype=float32
    """
    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    data = rows[1:]  # 跳过表头
    n = len(data)
    ax = np.empty(n, dtype=np.float32)
    ay = np.empty(n, dtype=np.float32)
    az = np.empty(n, dtype=np.float32)
    for i, row in enumerate(data):
        ax[i] = row[2]
        ay[i] = row[3]
        az[i] = row[4]
    wb.close()
    return ax, ay, az


def _fix_nan(arr: np.ndarray, name: str) -> np.ndarray:
    """用线性插值修复数组中的 NaN 值"""
    nan_count = np.isnan(arr).sum()
    if nan_count == 0:
        return arr
    print(f"  {name}: 发现 {nan_count} 个 NaN，线性插值修复")
    idx = np.arange(len(arr))
    good = ~np.isnan(arr)
    arr = np.interp(idx, idx[good], arr[good]).astype(np.float32)
    return arr


def load_csv(path: str | Path) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    读取 TCP receiver 生成的 CSV 文件中的 3 轴加速度数据。
    列布局: timestamp_us, datetime, acc_x, acc_y, acc_z
    返回 (ax, ay, az), dtype=float32
    """
    ax_list, ay_list, az_list = [], [], []
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)  # 跳过表头
        for row in reader:
            ax_list.append(float(row[2]))
            ay_list.append(float(row[3]))
            az_list.append(float(row[4]))
    ax = _fix_nan(np.array(ax_list, dtype=np.float32), "X")
    ay = _fix_nan(np.array(ay_list, dtype=np.float32), "Y")
    az = _fix_nan(np.array(az_list, dtype=np.float32), "Z")
    return ax, ay, az


def load_data(path: str | Path) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    自动识别 xlsx/csv 并加载 3 轴加速度数据。
    返回 (ax, ay, az), dtype=float32
    """
    path = Path(path)
    if path.suffix == '.csv':
        return load_csv(path)
    elif path.suffix == '.xlsx':
        return load_xlsx(path)
    else:
        raise ValueError(f"不支持的文件格式: {path.suffix}，仅支持 .xlsx 和 .csv")


def remove_dc_offset(
    ax: np.ndarray,
    ay: np.ndarray,
    az: np.ndarray,
    static_n: int = 10000,
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    用前 static_n 行静态数据计算直流偏移，减去后返回交流分量。
    """
    dc_x = ax[:static_n].mean()
    dc_y = ay[:static_n].mean()
    dc_z = az[:static_n].mean()
    print(f"直流偏移 (前{static_n}行): X={dc_x:.6f}, Y={dc_y:.6f}, Z={dc_z:.6f}")
    return ax - dc_x, ay - dc_y, az - dc_z
